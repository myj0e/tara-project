import re
import json
import os
import hashlib
import pymysql
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from pymysql.cursors import DictCursor

#mysql配置
db_config = {
    'host': '',   #IP
    'user': 'root',
    'password': '123456',
    'database': '',
    'charset': 'utf8mb4'
}



#==============================================================================================================================
'''
    实现存取的函数部分
    实现存取的函数部分
'''
#最初DFD的存取操作
def putDFDJSON2db(hash_id, dfd_json):
    '''
    将 DFD JSON 写入数据库
    '''
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()
    # print(dfd_json)
    print(dfd_json['title'])
    try:
        print("插入INITmeta表")
        cursor.execute(
            '''
            INSERT INTO init_db.dfd_meta (hash_id,title,description)
            VALUES (%s, %s, %s)
            ''',
            (hash_id, dfd_json['title'], dfd_json['description'])
        )
        print("插入nodes信息")
        for node in dfd_json['nodes']:
            cursor.execute(
                """
                INSERT INTO init_db.nodes (hash_id,graph_id,name,description,type,has_open_threats)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (hash_id, node['id'], node['name'], node['description'],node['type'],
                 node.get('hasOpenThreats', False))

            )
            node_id=node['id']
            for threat in node.get('threats', []):
                cursor.execute(
                    '''
                    INSERT INTO init_db.threats (id,hash_id,title, severity, type, mitigation, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ''',
                    (
                        node_id,
                        hash_id,
                        threat['title'],
                        threat['severity'],
                        threat['type'],
                        threat['mitigation'],
                        threat['status']
                    )
                )
        for edge in dfd_json['edges']:
            cursor.execute(
                '''
                INSERT INTO init_db.edges
                (hash_id, edge_id,name,description, source, target,
                 is_encrypted, is_public_network, has_open_threats, protocol)
                VALUES (%s, %s,%s , %s, %s, %s, %s, %s, %s, %s)
                ''',
                (
                    hash_id,
                    edge['id'],
                    edge['name'],
                    edge['description'],
                    edge['source'],
                    edge['target'],
                    edge.get('isEncrypted', False),
                    edge.get('isPublicNetwork', False),
                    edge.get('hasOpenThreats', False),
                    edge.get('protocol', '')
                )
            )
            # 获取刚插入的边ID
            edge_id = edge['id']
            for threat in edge.get('threats', []):
                cursor.execute(
                    '''
                    INSERT INTO init_db.threats (id,hash_id, title, severity, type, mitigation, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ''',
                    (
                        edge_id,
                        hash_id,
                        threat['title'],
                        threat['severity'],
                        threat['type'],
                        threat['mitigation'],
                        threat['status']
                    )
                )
        print(f"DFD JSON 写入成功，hash_id={hash_id}")
        connection.commit()

    except Exception as e:
        connection.rollback()
        print(hash_id+"写入 def_json 失败:", e)
    finally:
        connection.close()
        cursor.close()
def getDFDJSON(hash_id):
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor(pymysql.cursors.DictCursor)
    try:
        # 1) 节点明细 + 其威胁
        cursor.execute("""
            SELECT node.graph_id, node.name, node.description, node.type, node.has_open_threats,
                   threat.title, threat.severity, threat.type AS threat_type, threat.mitigation, threat.status
            FROM init_db.nodes AS node
            LEFT JOIN init_db.threats AS threat
                   ON threat.id = node.graph_id
            WHERE node.hash_id = %s AND node.graph_id IS NOT NULL
            ORDER BY node.id
        """, (hash_id,))
        node_rows = cursor.fetchall()

        # 2) 边明细 + 其威胁
        cursor.execute("""
            SELECT edge.edge_id, edge.name, edge.description, edge.source, edge.target,
                   edge.is_encrypted, edge.is_public_network, edge.has_open_threats, edge.protocol,
                   threat.title, threat.severity, threat.type AS threat_type, threat.mitigation, threat.status
            FROM init_db.edges AS edge
            LEFT JOIN init_db.threats AS threat
                   ON threat.id = edge.edge_id
            WHERE edge.hash_id = %s
            ORDER BY edge.id
        """, (hash_id,))
        edge_rows = cursor.fetchall()

        # 3) 标题与描述
        cursor.execute("""
            SELECT title, description
            FROM init_db.dfd_meta
            WHERE hash_id = %s
        """, (hash_id,))
        title_row = cursor.fetchone()

        # 4) 聚合节点
        nodes_dict = {}
        for row in node_rows:
            node_id = row['graph_id']
            if node_id not in nodes_dict:
                nodes_dict[node_id] = {
                    'id': node_id,
                    'name': row['name'],
                    'description': row['description'] or '',
                    'type': row['type'],
                    'hasOpenThreats': bool(row['has_open_threats']),
                    'threats': []
                }
            # 只有当 JOIN 出来的威胁不为空才追加
            if row['title'] is not None:
                nodes_dict[node_id]['threats'].append({
                    'title': row['title'],
                    'severity': row['severity'],
                    'type': row['threat_type'],
                    'mitigation': row['mitigation'],
                    'status': row['status']
                })

        # 5) 聚合边
        edges_dict = {}
        for row in edge_rows:
            edge_id = row['edge_id']
            if edge_id not in edges_dict:
                edges_dict[edge_id] = {
                    'id': edge_id,
                    'name': row['name'],
                    'description': row['description'] or '',
                    'source': row['source'],
                    'target': row['target'],
                    'isEncrypted': bool(row['is_encrypted']),
                    'isPublicNetwork': bool(row['is_public_network']),
                    'hasOpenThreats': bool(row['has_open_threats']),
                    'protocol': row['protocol'] or '',
                    'threats': []
                }
            if row['title'] is not None:
                edges_dict[edge_id]['threats'].append({
                    'title': row['title'],
                    'severity': row['severity'],
                    'type': row['threat_type'],
                    'mitigation': row['mitigation'],
                    'status': row['status']
                })

        # 6) 组装最终返回
        result = {
            'title': (title_row['title'] if title_row else ''),
            'description': (title_row['description'] if title_row else ''),
            'nodes': list(nodes_dict.values()),
            'edges': list(edges_dict.values())
        }
        return result

    except Exception as e:
        print("getDFDJSON读取数据库失败:", e)
        return {}
    finally:
        cursor.close()
        connection.close()
def getTitle(hash_id):
    connection = pymysql.connect(**db_config)
    cursor=connection.cursor()
    try:
        cursor.execute('''
        SELECT title FROM init_db.dfd_meta WHERE hash_id = %s
        ''',(hash_id,)
        )
        result = cursor.fetchone()
        if result:
            return result
        else:
            return None
        connection.commit()
    except Exception as e:
        print("获取标题名称失败:", e)
        return None
    finally:
        connection.close()
        cursor.close()

#Dread的存取操作
def putDreadJSON2db(hash_id, dread_json):
    """
    将 dread_json 写入 dread_db
    """
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()
    try:
        for node in dread_json.get('nodes', []):
            # 插入节点
            cursor.execute(
                """
                INSERT INTO dread_db.nodes (hash_id, name, type)
                VALUES (%s, %s, %s)
                """,
                (hash_id, node['name'], node['type'])
            )
            node_id = cursor.lastrowid

            # 插入该节点的 stride 威胁
            for stride_type, stride_info in node.get('stride', {}).items():
                dread_score = stride_info.get('dread', {})
                cursor.execute(
                    """
                    INSERT INTO dread_db.stride (
                        hash_id, node_id, stride_type,
                        description, scenario,
                        dread_D, dread_R, dread_E, dread_A, dread_D2
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        hash_id, node_id, stride_type,
                        stride_info.get('description', ''),
                        stride_info.get('Scenario', ''),
                        dread_score.get('D', 0),
                        dread_score.get('R', 0),
                        dread_score.get('E', 0),
                        dread_score.get('A', 0),
                        dread_score.get('D2', 0)
                    )
                )

        connection.commit()
        print(f"dread_json 写入成功 hash_id={hash_id}")

    except Exception as e:
        connection.rollback()
        print("写入 dread_json 失败:", e)
    finally:
        cursor.close()
        connection.close()

def getDreadJSON(hash_id):
    """
    从 MySQL dread_db 读取数据并拼装为 JSON 字典
    最终返回一个JSON字典
    """
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor(pymysql.cursors.DictCursor)
    try:
        #1.查所有节点
        cursor.execute("SELECT id, name, type FROM dread_db.nodes WHERE hash_id= %s ORDER BY id", (hash_id,))
        nodes = cursor.fetchall()

        #2.查stride 记录
        cursor.execute("""
            SELECT node_id, stride_type, description, scenario,
                   dread_D, dread_R, dread_E, dread_A, dread_D2
            FROM dread_db.stride
            WHERE hash_id = %s
            ORDER BY id
        """, (hash_id,))
        stride_records = cursor.fetchall()

        #3.将 stride 按 node_id 分组
        stride_map = {}
        for s in stride_records:
            node_id = s['node_id']
            # 构造每条stride对应的JSON
            stride_entry = {
                "description": s['description'] or "None",
                "Scenario": s['scenario'] or "None",
                "dread": {
                    "D": s['dread_D'],
                    "R": s['dread_R'],
                    "E": s['dread_E'],
                    "A": s['dread_A'],
                    "D2": s['dread_D2']
                }
            }
            # stride_type 是 S/T/R/I/D/E
            stride_map.setdefault(node_id, {})[s['stride_type']] = stride_entry

        #4.拼装最终 JSON
        result = {"nodes": []}
        for node in nodes:
            result["nodes"].append({
                "name": node['name'],
                "type": node['type'],
                "stride": stride_map.get(node['id'], {})  #没有记录时返回空字典
            })
        return result
    except Exception as e:
        print("getDreadJSON读取数据库失败:", e)
        return {}
    finally:
        cursor.close()
        connection.close()

#ThreadModleJSON的存取操作
def putThreatModleJSON2db(hash_id,threat_model_json):
    '''
    将threat_model_json写入 数据库 中
    '''
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()
    try:
        for threat in threat_model_json['nodes']:
            cursor.execute(
                """
                INSERT INTO threat_models_db.nodes (hash_id ,name ,type)
                VALUES (%s, %s, %s)
                """,
                (hash_id, threat['name'], threat['type'])
            )
            # 获取刚插入的节点ID
            node_id= cursor.lastrowid
            for stride in threat['stride']:
                cursor.execute(
                    """
                    INSERT INTO threat_models_db.stride_threats (hash_id, node_id, threat_type, scenario, potential_impact)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (hash_id, node_id, stride['Threat Type'], stride['Scenario'], stride['Potential Impact'])
                )
        # for suggestion in threat_model_json['improvement_suggestions']:
        #     cursor.execute(
        #         """
        #         INSERT INTO threat_models_db.improvement_suggestions (hash_id, suggestion)
        #         VALUES (%s, %s)
        #         """,
        #         (hash_id, suggestion)
        #     )
        connection.commit()
    except Exception as e:
        connection.rollback()
        print(f"写入threat_model_json失败: ", e)
    finally:
        cursor.close()
        connection.close()
def getThreatModleJSON(hash_id):
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor(pymysql.cursors.DictCursor)  # 返回字典
    try:
        #1.查询所有节点
        cursor.execute("SELECT id, name, type FROM threat_models_db.nodes WHERE hash_id = %s", (hash_id,))
        nodes = cursor.fetchall()

        #2.查询所有 stride 记录
        cursor.execute("""
            SELECT threat_id,node_id, threat_type, scenario, potential_impact
            FROM threat_models_db.stride_threats
            WHERE hash_id = %s
            ORDER BY threat_id
        """, (hash_id,))
        stride_records = cursor.fetchall()

        #3.将 stride 按 node_id 分组
        stride_map = {}
        for s in stride_records:
            node_id = s['node_id']
            stride_entry = {
                "Threat Type": s['threat_type'],
                "Scenario": s['scenario'] or "None",
                "Potential Impact": s['potential_impact'] or "None"
            }
            stride_map.setdefault(node_id, []).append(stride_entry)

        # 4.查询改进建议
        # cursor.execute("""
        #     SELECT suggestion FROM threat_models_db.improvement_suggestions
        #     WHERE hash_id = %s
        # """, (hash_id,))
        # suggestions = [row['suggestion'] for row in cursor.fetchall()]

        #5.拼装最终 JSON
        result = {
            "nodes": [],
            # "improvement_suggestions": suggestions
        }
        for node in nodes:
            result["nodes"].append({
                "name": node['name'],
                "type": node['type'],
                "stride": stride_map.get(node['id'], [])  #没有记录时返回空列表
            })

        return result

    except Exception as e:
        print("getThreatModleJSON读取数据库失败:", e)
        return {}
    finally:
        cursor.close()
        connection.close()

# CommitJSON的存取操作
def putCommitJSON2db(hash_id, commit_json):
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()
    try:
        # ---------- Step 1: 先清空旧数据 ----------
        # 若存在外键约束：先删子表再删父表
        cursor.execute(
            "DELETE FROM commit_db.stride WHERE hash_id = %s",
            (hash_id,)
        )
        cursor.execute(
            "DELETE FROM commit_db.nodes WHERE hash_id = %s",
            (hash_id,)
        )

        # ---------- Step 2: 重新插入数据 ----------
        for node in commit_json.get('nodes', []):
            node_name = (node.get('name') or '').strip()
            node_type = (node.get('type') or '').strip()

            cursor.execute(
                """
                INSERT INTO commit_db.nodes (hash_id, name, type)
                VALUES (%s, %s, %s)
                """,
                (hash_id, node_name, node_type)
            )
            node_id = cursor.lastrowid

            # 遍历 stride 字典
            for stride_type, stride_info in (node.get('stride') or {}).items():
                # 防御式处理，避免 None
                description = (stride_info.get('description') or '').strip()
                scenario = (stride_info.get('Scenario') or '').strip()
                dread = stride_info.get('dread') or {}
                commit_msg = (stride_info.get('commit') or '').strip()

                # 约束到整数（0-10），如果你的评分是 0-10
                def _score(x):
                    try:
                        v = int(x)
                    except Exception:
                        v = 0
                    return max(0, min(10, v))

                D = _score(dread.get('D', 0))
                R = _score(dread.get('R', 0))
                E = _score(dread.get('E', 0))
                A = _score(dread.get('A', 0))
                D2 = _score(dread.get('D2', 0))

                cursor.execute(
                    """
                    INSERT INTO commit_db.stride (
                        hash_id, node_id, stride_type,
                        description, scenario,
                        D, R, E, A, D2, commit_msg
                    )
                    VALUES (%s,%s,%s, %s,%s, %s,%s,%s,%s,%s,%s)
                    """,
                    (
                        hash_id, node_id, stride_type,
                        description, scenario,
                        D, R, E, A, D2, commit_msg
                    )
                )

        # ---------- Step 3: 提交 ----------
        connection.commit()
        print(f"[putCommitJSON2db] 写入成功，hash_id={hash_id}")

    except Exception as e:
        connection.rollback()
        print(f"[putCommitJSON2db] 写入失败（已回滚）：{e}")
        raise
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            connection.close()
        except Exception:
            pass
def getCommitJSON(hash_id):
    """
        从 commit_db 导出指定 hash_id 的 JSON 字典
        返回格式与示例一致
        """
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor(pymysql.cursors.DictCursor)
    try:
        #1.查询节点
        cursor.execute(
            "SELECT id,name, type FROM commit_db.nodes WHERE hash_id=%s ORDER BY id",
            (hash_id,)
        )
        nodes = cursor.fetchall()

        #2.查询对应的 stride 威胁信息
        cursor.execute("""
                SELECT node_id, stride_type, description, scenario,
                       D, R, E, A, D2, commit_msg
                FROM commit_db.stride
                WHERE hash_id=%s
                ORDER BY id
            """, (hash_id,))
        stride_records = cursor.fetchall()

        #3.将 stride 记录按 node_id 分组
        stride_map = {}
        for s in stride_records:
            node_id = s['node_id']
            stride_entry = {
                "description": s['description'] or "None",
                "Scenario": s['scenario'] or "None",
                "dread": {
                    "D": s['D'],
                    "R": s['R'],
                    "E": s['E'],
                    "A": s['A'],
                    "D2": s['D2']
                }
            }
            stride_entry['commit'] = s['commit_msg']

            stride_map.setdefault(node_id, {})[s['stride_type']] = stride_entry

        #4.组装 JSON 字典
        result = {"nodes": []}
        for node in nodes:
            result["nodes"].append({
                "name": node['name'],
                "type": node['type'],
                "stride": stride_map.get(node['id'], {})  #没有威胁就空字典
            })

        return result

    except Exception as e:
        print("导出 JSON 失败:", e)
        return {}
    finally:
        cursor.close()
        connection.close()

#消减措施存取
def putMitigationJSON2db(hash_id, mitigation_json):
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()
    # dfd_title：优先用传入JSON里带的；没有就回退到getTitle()
    title=getTitle(hash_id)

    summary_field = mitigation_json.get("summary", "")
    if isinstance(summary_field, list):
        summary_text = "；".join(s.strip() for s in summary_field if str(s).strip())
    else:
        summary_text = str(summary_field or "").strip()

    try:
        connection.begin()

        #插一次头表”，循环外；幂等：重复就更新标题并拿到现有id
        cursor.execute("""
            INSERT INTO mitigation_db.simple_mitigations (hash_id, dfd_title)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE
                id = LAST_INSERT_ID(id),      -- 让 lastrowid 指向已存在那行
                dfd_title = VALUES(dfd_title) -- 如不想更新标题，可改成 id = LAST_INSERT_ID(id)
        """, (hash_id, title))
        simple_id = cursor.lastrowid

        # 幂等写入——先清掉该hash_id旧的明细，避免重复累积
        # cursor.execute("DELETE FROM mitigation_db.mitigations WHERE hash_id=%s", (hash_id,))

        # 再展开 items 插入明细
        for item in mitigation_json.get("items", []):
            component   = (item.get("component") or "").strip()
            threat_type = (item.get("threat_type") or "").strip()
            scenario    = (item.get("scenario") or "").strip()
            mitigations = item.get("mitigations") or []

            # mitigations是字符串，用；拆分
            if isinstance(mitigations, str):
                mitigations = [x.strip() for x in mitigations.split("；") if x.strip()]

            for m in mitigations:
                mitigation_text = str(m).strip()
                if not mitigation_text:
                    continue
                cursor.execute(
                    """
                    INSERT INTO mitigation_db.mitigations
                      (hash_id, dfd_title, component, threat_type, scenario, mitigation, summary)
                    VALUES
                      (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (hash_id, title, component, threat_type, scenario, mitigation_text, summary_text)
                )

        connection.commit()
        print(f"写入 mitigation_json 成功，hash_id={hash_id}, simple_id={simple_id}")

    except Exception as e:
        connection.rollback()
        print(f"写入 mitigation_json 失败: {e}")
        raise
    finally:
        cursor.close()
        connection.close()

def getMitigationJSON2db(hash_id):
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor(pymysql.cursors.DictCursor)
    try:
        cursor.execute('''
            SELECT
              dfd_title,
              component,
              threat_type,
              scenario,
              GROUP_CONCAT(mitigation ORDER BY id SEPARATOR '；') AS mitigations,
              MAX(summary) AS summary
            FROM mitigation_db.mitigations
            WHERE hash_id = %s
            GROUP BY dfd_title, component, threat_type, scenario
            ORDER BY component, threat_type, dfd_title
        ''', (hash_id,))
        rows = cursor.fetchall()

        if not rows:
            print(f"没有找到对应的 mitigation_json，hash_id={hash_id}")
            return None

        # 统一返回结构（便于直接转 Markdown 表或导出）
        result = {
            "dfd_title": rows[0]['dfd_title'],
            "hash_id": hash_id,
            "items": [
                {
                    "component": r['component'],
                    "threat_type": r['threat_type'],
                    "scenario": r['scenario'],
                    "mitigations": r['mitigations'] or ""
                }
                for r in rows
            ],
            "summary": rows[0]['summary'] or ""
        }
        return result

    except Exception as e:
        print(f"获取 mitigation_json 失败: {e}")
        raise
    finally:
        cursor.close()
        connection.close()

def getMitigationHashIdById(id):
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor(pymysql.cursors.DictCursor)
    print("进入查询")

    try:
        cursor.execute(
            '''
            SELECT 
                hash_id
            FROM mitigation_db.simple_mitigations
            WHERE id = %s
            ''',
            (id,)
        )
        row = cursor.fetchone()
        if not row or row.get('hash_id') is None:
            print(f"未找到 hash_id，id={id}")
            return None
        return row['hash_id']

    except Exception as e:
        print(f"获取 mitigation_json 失败: {e}")
        raise
    finally:
        cursor.close()
        connection.close()

#导出execle
def exportMitigationToExcel(id, file_path):
    if file_path is None:
        print(f"没有路径，无法导出")
        return

    hash_id = getMitigationHashIdById(id)
    data = getMitigationJSON2db(hash_id)

    if not data:
        print(f"没有数据，无法导出 id={id}")
        df = pd.DataFrame(["异常"])
        df.insert(0, "DFD标题", "异常，未查询到数据")
        df.insert(1, "ID", "异常，未查询到数据")
        df.to_excel(file_path, index=False)
        print(f"已导出缓解措施表到 {file_path}")
        return

    # 生成 DataFrame
    df = pd.DataFrame(data["items"])
    df.insert(0, "DFD标题", data["dfd_title"])
    df.insert(1, "Hash ID", data["hash_id"])
    # 不直接放 Summary 到每行
    # df["Summary"] = data["summary"]

    # 导出 Excel
    df.to_excel(file_path, index=False)

    # 用 openpyxl 打开再修改
    wb = load_workbook(file_path)
    ws = wb.active

    # 在第一行插入 Summary 横幅
    ws.insert_rows(1)
    total_cols = ws.max_column
    last_col_letter = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"Summary: {data['summary']}"

    # 样式美化（可选）
    from openpyxl.styles import Alignment, Font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].font = Font(bold=True, size=12)

    wb.save(file_path)
    print(f"已导出缓解措施表到 {file_path}，Summary 置顶横幅完成")
#分页查询
def get_mitigations_page(page: int, page_size: int):
    """
    分页获取 mitigation 简要信息
    :param page: 当前页码，从 1 开始
    :param page_size: 每页记录数
    :return: 查询结果列表
    """
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor(pymysql.cursors.DictCursor)
    try:
        offset = (page - 1) * page_size
        cursor.execute('''
            SELECT id, created_at, dfd_title
            FROM mitigation_db.simple_mitigations
            ORDER BY created_at DESC
            LIMIT %s OFFSET %s
        ''', (page_size, offset))
        rows = cursor.fetchall()
        return rows
    except Exception as e:
        print(f"分页查询失败: {e}")
        return []
    finally:
        cursor.close()
        connection.close()

#子查询优化
def get_mitigations_page_with_count(page: int = 1, page_size: int = 5):
    page = max(int(page or 1), 1)
    page_size = max(int(page_size or 10), 1)
    offset = (page - 1) * page_size

    conn = pymysql.connect(
        cursorclass=DictCursor,
        **db_config
    )
    # 或者：cur = conn.cursor(DictCursor)
    cur  = conn.cursor()
    print("进入子查询优化")
    try:
        # 1) 统计总数
        cur.execute("SELECT COUNT(*) AS total FROM mitigation_db.simple_mitigations")
        total = cur.fetchone()['total']

        # 2) 子查询 + 回表
        sql = """
            SELECT m.id, m.created_at, m.dfd_title
            FROM mitigation_db.simple_mitigations AS m
            JOIN (
                SELECT id
                FROM mitigation_db.simple_mitigations
                ORDER BY created_at DESC, id DESC
                LIMIT %s OFFSET %s
            ) AS s ON s.id = m.id
            ORDER BY m.created_at DESC, m.id DESC
        """
        cur.execute(sql, (page_size, offset))
        rows = cur.fetchall()

        return {"total": total, "data": rows, "page": page, "page_size": page_size}
    finally:
        cur.close()
        conn.close()

def findMitigationsMdbyHashID(hash_id: str, db_cfg=None) -> bool:
    """
    判断该 hash_id 是否已在 MySQL 写入过缓解措施数据。
    命中条件：
      - mitigation_db.mitigations 存在该 hash_id 的任意一行；或
      - mitigation_db.simple_mitigations 存在该 hash_id 的头表记录。
    返回:
      True  - 已存在
      False - 不存在或查询异常
    """
    conn = pymysql.connect(**db_config)
    cur = conn.cursor()
    try:
        # 1) 先查明细表（效率高，加 LIMIT 1）
        cur.execute(
            "SELECT 1 FROM mitigation_db.mitigations WHERE hash_id = %s LIMIT 1",
            (hash_id,)
        )
        if cur.fetchone():
            print(f"{hash_id} mitigation 明细已存在")
            return True

        # 2) 再查头表
        cur.execute(
            "SELECT 1 FROM mitigation_db.simple_mitigations WHERE hash_id = %s LIMIT 1",
            (hash_id,)
        )
        exists = cur.fetchone() is not None
        if exists:
            print(f"{hash_id} mitigation 头表已存在（无明细）")
        else:
            print(f"{hash_id} mitigation 未找到")
        return exists

    except Exception as e:
        print(f"查询 mitigation 失败: {e}")
        return False
    finally:
            if cur: cur.close()
            if conn: conn.close()

#==============================================================================================================================
'''
    通过hash_id查询有没有的函数
'''
def findDFDJSONbyHashID(hash_id):
    '''
    在DFD JSON表中查 存不存在指定的id；返回True or False
    '''
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()
    try:
        cursor.execute(
            '''
            SELECT 1 FROM init_db.nodes WHERE hash_id = %s LIMIT 1
            '''
            ,(hash_id,)
        )
        result = cursor.fetchone()
        if(result):
            print(f"DFD JSON 存在，hash_id={hash_id}")
            return True
        else :
            print(f"DFD JSON 不存在，hash_id={hash_id}")
            return False
        connection.commit();
    except Exception as e:
        print("查询 DFD JSON 失败:", e)
        return False
    finally:
        connection.close();
        cursor.close();

def findDreadJSONbyHashID(hash_id):
    '''
    在Dread JSON表中查 存不存在指定的id；返回True or False
    '''
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()
    try:
        cursor.execute(
            '''
            SELECT 1 FROM dread_db.nodes WHERE hash_id = %s LIMIT 1
            '''
            ,(hash_id,)
        )
        result = cursor.fetchone()
        if(result):
            print(f"Dread JSON 存在，hash_id={hash_id}")
            return True
        else :
            print(f"Dread JSON 不存在，hash_id={hash_id}")
            return False
        connection.commit();
    except Exception as e:
        print("查询 Dread JSON 失败:", e)
        return False
    finally:
        connection.close();
        cursor.close();

def findThreatModelsJSONbyHashID(hash_id):
    '''
    ThreatModels JSON表中查 存不存在指定的id；返回True or False
    '''
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()
    try:
        cursor.execute(
            '''
            SELECT 1 FROM threat_models_db.nodes WHERE hash_id = %s LIMIT 1
            '''
            ,(hash_id,)
        )
        result = cursor.fetchone()
        if(result):
            print(f"ThreatModels JSON 存在，hash_id={hash_id}")
            return True
        else :
            print(f"ThreatModels JSON 不存在，hash_id={hash_id}")
            return False
        connection.commit();
    except Exception as e:
        print("查询 ThreatModels JSON 失败:", e)
        return False
    finally:
        connection.close();
        cursor.close();

def findCommitJSONbyHashID(hash_id):
    '''
    Commit JSON表中查 存不存在指定的id；返回True or False
    '''
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()
    try:
        cursor.execute(
            '''
            SELECT 1 FROM commit_db.nodes WHERE hash_id = %s LIMIT 1
            '''
            ,(hash_id,)
        )
        result = cursor.fetchone()
        if(result):
            print(f"Commit JSON 存在，hash_id={hash_id}")
            return True
        else :
            print(f"Commit JSON 不存在，hash_id={hash_id}")
            return False
        connection.commit();
    except Exception as e:
        print("查询 DFD JSON 失败:", e)
        return False
    finally:
        connection.close();
        cursor.close();

#==============================================================================================================================
